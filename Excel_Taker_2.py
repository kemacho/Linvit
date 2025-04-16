import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import openpyxl
import os
from datetime import datetime
import threading
from tkinterdnd2 import DND_FILES, TkinterDnD
import shutil
from string import ascii_uppercase as alc

# Путь к файлу-шаблону
TEMPLATE_FILE_PATH = r'\\192.168.34.9\линвит\ПОЛЬЗОВАТЕЛИ\USER49\Программы\Шаблоны для программ\Файл для экспорта (для ИК).xlsx'

# Цветовая схема
BG_COLOR = "#f0f0f0"
BUTTON_COLOR = "#4b8fe2"
BUTTON_HOVER = "#3a7bc8"
TEXT_COLOR = "#333333"
FRAME_COLOR = "#ffffff"
LISTBOX_COLOR = "#ffffff"
PROGRESS_COLOR = "#4b8fe2"

# Функция для форматирования даты
def format_date(date_value):
    if isinstance(date_value, datetime):
        return date_value.strftime("%d.%m.%Y")
    else:
        return str(date_value or "")


def process_files(action):
    """
    Обрабатывает выбранные файлы, используя заранее известный индекс листа.
    """
    global cell_value1, cell_value2, cell_value3, cell_value4, cell_value4_1, cell_value5, cell_value6, cell_value7, cell_value8, cell_value10, cell_value10_1, cell_value9, cell_value12, cell_value13, cell_value14, cell_value11, cell_value15, cell_value16, cell_value17
    update_status("Начало обработки файлов...")  # Обновление статуса

    if not file_paths:
        messagebox.showinfo("Внимание", "Файлы не выбраны.")
        update_status("Ожидание выбора файлов.")
        return

    if action == "existing":
        save_path = filedialog.askopenfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not save_path:
            update_status("Обработка прервана пользователем.")
            return

        try:
            output_workbook = openpyxl.load_workbook(save_path)
            output_sheet = output_workbook.worksheets[0]
        except FileNotFoundError:
            messagebox.showerror("Ошибка", f"Файл не найден: {save_path}")
            update_status(f"Ошибка: Файл не найден - {save_path}")
            return
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при открытии файла {save_path}: {e}")
            update_status(f"Ошибка: Не удалось открыть файл - {save_path}. Ошибка: {e}")
            return

    elif action == "new":
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not save_path:
            update_status("Обработка прервана пользователем.")
            return

        # Копируем файл шаблона
        try:
            shutil.copy(TEMPLATE_FILE_PATH, save_path)
            output_workbook = openpyxl.load_workbook(save_path)
            output_sheet = output_workbook.worksheets[0]
        except FileNotFoundError:
            messagebox.showerror("Ошибка", f"Файл шаблона не найден: {TEMPLATE_FILE_PATH}")
            update_status(f"Ошибка: Файл шаблона не найден - {TEMPLATE_FILE_PATH}")
            return
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при копировании шаблона: {e}")
            update_status(f"Ошибка: Не удалось скопировать шаблон - {TEMPLATE_FILE_PATH}. Ошибка: {e}")
            return

    else:
        messagebox.showerror("Ошибка", "Некорректное действие.")
        update_status("Ошибка: Некорректное действие.")
        return

    row_index = output_sheet.max_row + 1
    total_files = len(file_paths)
    progress_bar['maximum'] = total_files
    progress_bar['value'] = 0

    for i, file_path in enumerate(file_paths):
        try:
            update_status(f"Обработка файла {i + 1} из {total_files}: {os.path.basename(file_path)}")

            workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            worksheets = workbook.sheetnames

            sheet1 = workbook['Титул']
            sheet2 = workbook['Протокол']
            sheet3 = workbook['Записи']

            for i in range(0, len(worksheets)):
                if worksheets[i] == 'Протокол-3пр':
                    sheet2 = workbook['Протокол-3пр']
            for i in range(0, len(worksheets)):
                if worksheets[i] == 'Записи-3пр':
                    sheet3 = workbook['Записи-3пр']

            # Поиск ячейки с "отрицательное отклонение напряжения"
            cell_value16 = None  # Значение по умолчанию
            cell_value17 = None
            search_text = "отрицательное отклонение напряжения"

            for row in range(70, 90):  # Поиск в строках
                for col in range(8, 10):  # Проверяем столбцы
                    cell = sheet2.cell(row=row, column=col)
                    if cell.value and search_text in str(cell.value):
                        # Нашли ячейку
                        cell_value16 = sheet2.cell(row=row + 1, column=11).value # dU -
                        cell_value17 = sheet2.cell(row=row + 3, column=11).value # dU +
                        break
                if cell_value16 is not None:
                    break

            cell_value8 = sheet3['AK6'].value  # Место в схеме
            cell_value13 = sheet3['U9'].value  # Центр питания

            if cell_value8 is None:
                cell_value8 = '-'
            if cell_value13 is None:
                cell_value13 = '-'

            Check1 = str(sheet2['AG34'].value)
            Check2 = str(sheet2['AG32'].value)
            Check3 = str(sheet2['AG30'].value)
            Check4 = str(sheet2['AG33'].value)


            if 'Тип СИ' in Check4:
                cell_value2 = sheet2['AG34'].value  # Тип СИ ПКЭ
                cell_value1 = sheet2['BE34'].value  # Заводской номер СИ ПКЭ
                cell_value14 = sheet2['CD34'].value  # Поверка СИ ПКЭ

                cell_value10 = sheet2['AG35'].value  # Тип СИ
                cell_value11 = sheet2['BE35'].value  # Заводской номер СИ
                cell_value15 = sheet2['CD35'].value  # Поверка СИ

                cell_value4 = sheet1['A32'].value  # Электрические сети
                cell_value4_1 = sheet1['A33'].value  # Электрические сети
                cell_value9 = sheet1['BC26'].value  # Номер протокола
                cell_value12 = format_date(sheet1['BU24'].value)  # Дата протокола

                cell_value5 = format_date(sheet2['M25'].value)  # Начало испытаний
                cell_value6 = format_date(sheet2['M26'].value)  # Окончание испытаний


            elif 'Тип СИ' in Check3:
                cell_value2 = sheet2['AG31'].value  # Тип СИ ПКЭ
                cell_value1 = sheet2['BE31'].value  # Заводской номер СИ ПКЭ
                cell_value14 = sheet2['CD31'].value  # Поверка СИ ПКЭ

                cell_value10 = sheet2['AG32'].value  # Тип СИ
                cell_value11 = sheet2['BE32'].value  # Заводской номер СИ
                cell_value15 = sheet2['CD32'].value  # Поверка СИ

                cell_value4 = sheet1['A35'].value  # Электрические сети
                cell_value4_1 = sheet1['A36'].value  # Электрические сети
                cell_value9 = sheet1['BC29'].value  # Номер протокола
                cell_value12 = format_date(sheet1['BU24'].value)  # Дата протокола

                cell_value5 = format_date(sheet2['M22'].value)  # Начало испытаний
                cell_value6 = format_date(sheet2['M23'].value)  # Окончание испытаний



            elif 'Тип СИ' in Check2:
                cell_value2 = sheet2['AG33'].value  # Тип СИ ПКЭ
                cell_value1 = sheet2['BE33'].value  # Заводской номер СИ ПКЭ
                cell_value14 = sheet2['CD33'].value  # Поверка СИ ПКЭ

                cell_value10 = sheet2['AG34'].value  # Тип СИ
                cell_value11 = sheet2['BE34'].value  # Заводской номер СИ
                cell_value15 = sheet2['CD34'].value  # Поверка СИ

                cell_value4 = sheet1['A35'].value  # Электрические сети
                cell_value4_1 = sheet1['A36'].value  # Электрические сети
                cell_value9 = sheet1['BC29'].value  # Номер протокола
                cell_value12 = format_date(sheet1['BU24'].value)  # Дата протокола

                cell_value5 = format_date(sheet2['M24'].value)  # Начало испытаний
                cell_value6 = format_date(sheet2['M25'].value)  # Окончание испытаний


            elif 'Тип СИ' in Check1:
                cell_value2 = sheet2['AG35'].value  # Тип СИ ПКЭ
                cell_value1 = sheet2['BE35'].value  # Заводской номер СИ ПКЭ
                cell_value14 = sheet2['CD35'].value  # Поверка СИ ПКЭ

                cell_value10 = sheet2['AG36'].value  # Тип СИ
                cell_value11 = sheet2['BE36'].value  # Заводской номер СИ
                cell_value15 = sheet2['CD36'].value  # Поверка СИ

                cell_value4 = sheet1['A35'].value  # Электрические сети
                cell_value4_1 = sheet1['A36'].value  # Электрические сети
                cell_value9 = sheet1['BC29'].value  # Номер протокола
                cell_value12 = format_date(sheet1['BU24'].value)  # Дата протокола

                cell_value5 = format_date(sheet2['M26'].value)  # Начало испытаний
                cell_value6 = format_date(sheet2['M27'].value)  # Окончание испытаний


            if cell_value4_1 is not None:
                cell_value4 = str(cell_value4) + ' ' + str(cell_value4_1)

            if cell_value9 is not None:
                Protocol_Num0 = cell_value9.split('/')
                Protocol_Num = Protocol_Num0[0] + ',' + Protocol_Num0[1]
            else:
                Protocol_Num = '-'

            # Запись данных в первый лист
            output_sheet[f'B{row_index}'] = cell_value4  # Электрические сети
            output_sheet[f'C{row_index}'] = cell_value9  # Номер протокола
            output_sheet[f'D{row_index}'] = cell_value12  # Дата протокола
            output_sheet[f'E{row_index}'] = cell_value13  # Центр питания
            output_sheet[f'F{row_index}'] = cell_value8  # Место в схеме
            output_sheet[f'G{row_index}'] = cell_value5  # Дата начала испытаний
            output_sheet[f'H{row_index}'] = cell_value6  # Дата окончания испытаний
            output_sheet[f'I{row_index}'] = cell_value2  # Тип СИ ПКЭ
            output_sheet[f'J{row_index}'] = cell_value1  # Заводской № ПКЭ
            output_sheet[f'K{row_index}'] = cell_value14  # Поверка СИ ПКЭ
            output_sheet[f'L{row_index}'] = cell_value10  # Тип СИ
            output_sheet[f'M{row_index}'] = cell_value11  # Заводской № СИ
            output_sheet[f'N{row_index}'] = cell_value15  # Поверка СИ
            output_sheet[f'O{row_index}'] = cell_value16  # dU(-)
            output_sheet[f'P{row_index}'] = cell_value17  # dU(+)
            output_sheet[f'R{row_index}'] = Protocol_Num  # Порядковый номер протокола

            row_index += 1

        except FileNotFoundError:
            messagebox.showerror("Ошибка", f"Файл не найден: {file_path}")
            update_status(f"Ошибка: Файл не найден - {os.path.basename(file_path)}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при обработке файла {os.path.basename(file_path)}: {e}")
            update_status(f"Ошибка при обработке {os.path.basename(file_path)}: {e}")

        # Обновление прогресс-бара (основной поток)
        root.after(0, update_progress)

    try:
        update_status("Сохранение файла...")
        output_workbook.save(save_path)
        messagebox.showinfo("Успех", f"Данные успешно записаны в файл: {save_path}")
        update_status(f"Данные успешно записаны в файл: {save_path}")

    except Exception as e:
        messagebox.showerror("Ошибка", f"Ошибка при сохранении файла: {e}")
        update_status(f"Ошибка при сохранении файла: {e}")
    finally:
        output_workbook.close()
        update_status("Обработка завершена.")

    root.after(0, processing_complete)  # Сообщение о завершении


def update_progress():
    """Обновляет значение прогресс-бара."""
    progress_bar['value'] += 1


def processing_complete():
    """Сбрасывает прогресс-бар после завершения."""
    progress_bar['value'] = 0

def update_status(message):
    """Обновляет текст в метке статуса."""
    status_label.config(text=message)
    root.update_idletasks()  # Обновление интерфейса для отображения сообщения


def choose_files():
    """Выбор файлов."""
    global file_paths
    new_file_paths = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx")])
    if new_file_paths:
        file_paths.extend(new_file_paths)
        update_file_list()
    update_status("Файлы выбраны.")


def delete_selected_file():
    """Удаление выбранного файла."""
    try:
        selected_index = file_list.curselection()[0]
        del file_paths[selected_index]
        update_file_list()
    except IndexError:
        messagebox.showinfo("Внимание", "Файл не выбран для удаления.")
    update_status("Файл удален из списка.")


def clear_file_list():
    """Очистка списка файлов."""
    global file_paths
    file_paths = []
    update_file_list()
    update_status("Список файлов очищен.")


def update_file_list():
    """Обновление списка файлов."""
    file_list.delete(0, tk.END)
    for file_path in file_paths:
        file_list.insert(tk.END, os.path.basename(file_path))


def drop(event):
    """Обработчик события Drop."""
    global file_paths
    files = root.tk.splitlist(event.data)
    file_paths.extend(files)
    update_file_list()
    update_status("Файлы добавлены перетаскиванием.")


def start_processing_thread(action):
    """Запуск обработки в отдельном потоке."""
    thread = threading.Thread(target=process_files, args=(action,))
    thread.start()

def create_button(parent, text, command, width=20):
    """Создает стилизованную кнопку."""
    button = tk.Button(
        parent,
        text=text,
        command=command,
        bg=BUTTON_COLOR,
        fg="white",
        activebackground=BUTTON_HOVER,
        activeforeground="white",
        relief="flat",
        font=("Arial", 10, "bold"),
        padx=10,
        pady=5,
        width=width,
        borderwidth=0,
        highlightthickness=0
    )
    button.bind("<Enter>", lambda e: button.config(bg=BUTTON_HOVER))
    button.bind("<Leave>", lambda e: button.config(bg=BUTTON_COLOR))
    return button

# == UI Setup ==
root = TkinterDnD.Tk()  # Инициализируем TkinterDnD
root.title("Excel File Transfer")
root.geometry("700x700")  # Немного увеличил высоту окна
root.configure(bg=BG_COLOR)

# Стиль для рамок
style = ttk.Style()
style.configure("TFrame", background=BG_COLOR)
style.configure("TLabel", background=BG_COLOR, foreground=TEXT_COLOR, font=("Arial", 10))
style.configure("TButton", font=("Arial", 10, "bold"), padding=5)

# Главная рамка
main_frame = ttk.Frame(root, padding=20)
main_frame.pack(expand=True, fill="both")

# Заголовок
header_label = ttk.Label(
    main_frame,
    text="Обработка файлов Excel",
    font=("Arial", 14, "bold"),
    foreground=TEXT_COLOR
)
header_label.pack(pady=(0, 15))

# Рамка для выбора файлов - теперь занимает больше места
file_frame = ttk.LabelFrame(main_frame, text=" Выбор файлов ", padding=10)
file_frame.pack(fill="both", expand=True, pady=5)  # Изменил на fill="both" и expand=True

# Кнопка выбора файлов (сделал немного меньше)
choose_button = create_button(file_frame, "Выбрать файлы", choose_files, 20)  # Уменьшил ширину
choose_button.pack(pady=5)

# Подпись для области перетаскивания
drag_label = ttk.Label(file_frame, text="или перетащите файлы в область ниже:")
drag_label.pack()

# Список файлов - теперь больше по размеру
file_list = tk.Listbox(
    file_frame,
    height=12,  # Увеличил высоту
    bg=LISTBOX_COLOR,
    relief="solid",
    borderwidth=1,
    font=("Arial", 10)
)
file_list.pack(fill="both", expand=True, pady=5)  # Теперь расширяется во все стороны

# Привязка события Drop к Listbox
file_list.drop_target_register(DND_FILES)
file_list.dnd_bind('<<Drop>>', drop)

# Рамка для кнопок управления файлами (сделал кнопки меньше)
control_frame = ttk.Frame(file_frame)
control_frame.pack(fill="x", pady=5)

# Кнопки управления списком (уменьшил ширину)
delete_button = create_button(control_frame, "Удалить выбранный", delete_selected_file, 15)
delete_button.pack(side="left", padx=5, expand=True)

clear_button = create_button(control_frame, "Очистить список", clear_file_list, 15)
clear_button.pack(side="left", padx=5, expand=True)

# Рамка для действий (сделал меньше)
action_frame = ttk.LabelFrame(main_frame, text=" Действия ", padding=10)
action_frame.pack(fill="x", pady=5)

# Кнопки действий (сделал немного меньше)
new_button = create_button(action_frame, "Создать новую таблицу", lambda: start_processing_thread("new"), 20)
new_button.pack(pady=5)

existing_button = create_button(action_frame, "Добавить в существующую", lambda: start_processing_thread("existing"), 20)
existing_button.pack(pady=5)

# Рамка для прогресса (оставил как было)
progress_frame = ttk.LabelFrame(main_frame, text=" Прогресс ", padding=10)
progress_frame.pack(fill="x", pady=5)

# Остальной код остается без изменений...

# Прогресс-бар
progress_bar = ttk.Progressbar(
    progress_frame,
    orient="horizontal",
    length=100,
    mode="determinate",
    style="custom.Horizontal.TProgressbar"
)
style.configure("custom.Horizontal.TProgressbar", troughcolor=BG_COLOR, background=PROGRESS_COLOR)
progress_bar.pack(fill="x", pady=5)

# Метка статуса
status_label = ttk.Label(
    progress_frame,
    text="Ожидание выбора файлов.",
    anchor="center",
    font=("Arial", 9),
    wraplength=600
)
status_label.pack(fill="x", pady=5)

# Переменная списка файлов
file_paths = []

# Инициализация статуса
update_status("Ожидание выбора файлов.")

# Запуск UI
root.mainloop()