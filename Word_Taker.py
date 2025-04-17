import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinterdnd2 import TkinterDnD, DND_FILES
import docx
import openpyxl
from openpyxl.styles import Alignment
import os
import threading
import re



class WordToExcelConverter(TkinterDnD.Tk):
    def __init__(self):
        super().__init__()
        self.title("Word → Excel Converter")
        self.geometry("700x550")
        self.configure(bg="#f0f0f0")
        self.file_paths = []

        # Стили
        self.style = ttk.Style()
        self.style.configure("TButton", padding=6, font=('Segoe UI', 10))
        self.style.configure("TLabel", font=('Segoe UI', 9))
        self.style.configure("Title.TLabel", font=('Segoe UI', 12, 'bold'))

        self.create_widgets()

    def create_widgets(self):
        # Основные фреймы
        header_frame = ttk.Frame(self, padding=10)
        header_frame.pack(fill=tk.X)

        main_frame = ttk.Frame(self, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Заголовок
        ttk.Label(header_frame, text="Конвертер Word в Excel", style="Title.TLabel").pack()

        # Область файлов
        file_frame = ttk.LabelFrame(main_frame, text=" Файлы для обработки ", padding=10)
        file_frame.pack(fill=tk.BOTH, expand=True)

        # Список файлов с прокруткой
        self.file_list = tk.Listbox(
            file_frame,
            selectmode=tk.EXTENDED,
            height=12,
            bg="white",
            fg="#333",
            selectbackground="#4CAF50",
            selectforeground="white",
            font=('Segoe UI', 9),
            relief="flat",
            highlightthickness=0
        )
        self.file_list.pack(fill=tk.BOTH, expand=True, pady=5)

        # Настройка drag and drop
        self.file_list.drop_target_register(DND_FILES)
        self.file_list.dnd_bind('<<Drop>>', self.add_dropped_files)

        # Кнопки управления файлами
        btn_frame = ttk.Frame(file_frame)
        btn_frame.pack(fill=tk.X, pady=(5, 0))

        ttk.Button(btn_frame, text="Добавить", command=self.add_files).pack(side=tk.LEFT, fill=tk.X, expand=True,
                                                                            padx=2)
        ttk.Button(btn_frame, text="Удалить", command=self.remove_selected).pack(side=tk.LEFT, fill=tk.X, expand=True,
                                                                                 padx=2)
        ttk.Button(btn_frame, text="Очистить", command=self.clear_list).pack(side=tk.LEFT, fill=tk.X, expand=True,
                                                                             padx=2)

        # Кнопки обработки
        process_frame = ttk.Frame(main_frame)
        process_frame.pack(fill=tk.X, pady=(10, 0))

        ttk.Button(process_frame, text="Новая таблица",
                   command=lambda: self.start_processing("new")).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)
        ttk.Button(process_frame, text="Добавить в файл",
                   command=lambda: self.start_processing("existing")).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)

        # Прогресс и статус
        self.progress = ttk.Progressbar(main_frame, orient="horizontal", mode="determinate")
        self.progress.pack(fill=tk.X, pady=(15, 5))

        self.status = ttk.Label(main_frame, text="Готов к работе")
        self.status.pack(fill=tk.X)

    def add_dropped_files(self, event):
        files = self.tk.splitlist(event.data)
        added = 0
        for f in files:
            if f.lower().endswith(('.doc', '.docx')) and f not in self.file_paths:
                self.file_paths.append(f)
                added += 1
        self.update_file_list()
        self.update_status(f"Добавлено {added} файлов" if added else "Нет новых файлов для добавления")

    def add_files(self):
        files = filedialog.askopenfilenames(
            filetypes=[("Word Files", "*.docx;*.doc"), ("All files", "*.*")],
            title="Выберите файлы Word"
        )
        if files:
            for f in files:
                if f not in self.file_paths:
                    self.file_paths.append(f)
            self.update_file_list()
            self.update_status(f"Добавлено {len(files)} файлов")

    def remove_selected(self):
        selected = self.file_list.curselection()
        if selected:
            for i in reversed(selected):
                del self.file_paths[i]
            self.update_file_list()
            self.update_status(f"Удалено {len(selected)} файлов")
        else:
            self.update_status("Не выбраны файлы для удаления")

    def clear_list(self):
        if self.file_paths:
            self.file_paths = []
            self.update_file_list()
            self.update_status("Список файлов очищен")
        else:
            self.update_status("Список файлов уже пуст")

    def update_file_list(self):
        self.file_list.delete(0, tk.END)
        for f in self.file_paths:
            self.file_list.insert(tk.END, os.path.basename(f))

    def update_status(self, message):
        self.status.config(text=message)
        self.update_idletasks()

    def update_progress(self, value):
        self.progress['value'] = value
        self.update_idletasks()

    def start_processing(self, action):
        if not self.file_paths:
            messagebox.showwarning("Внимание", "Сначала выберите файлы!")
            return

        thread = threading.Thread(target=self.process_files, args=(action,), daemon=True)
        thread.start()


    def extract_data_from_word(self, filepath):
        try:
            doc = docx.Document(filepath)

            z_data = []
            i_data = []
            p_data = []


            for table in doc.tables:
                for row_idx, row in enumerate(table.rows):
                    row_text = [cell.text.strip() for cell in row.cells]

                    if row_idx == 5:
                        p_data.append(row_text[5])

                    # Проверяем каждую ячейку в строке на наличие ключевых слов
                    for cell_idx, cell in enumerate(row.cells):
                        text = cell.text.strip().lower()

                        if "заявитель" in text:
                            # Собираем 4 следующие строки из первого столбца
                            for j in range(1, 5):
                                if row_idx + j < len(table.rows):
                                    if table.rows[row_idx + j].cells:  # Проверяем, что есть ячейки
                                        z_data.append(table.rows[row_idx + j].cells[0].text.strip())

                        elif "изготовитель" in text:
                            # Собираем 4 следующие строки из первого столбца
                            for j in range(1, 5):
                                if row_idx + j < len(table.rows):
                                    if table.rows[row_idx + j].cells:  # Проверяем, что есть ячейки
                                        i_data.append(table.rows[row_idx + j].cells[0].text.strip())


            # Объединяем данные в строки с переносами
            z_result = " ".join(z_data[:4]) if z_data else ""
            i_result = " ".join(i_data[:4]) if i_data else ""
            p_result = p_data

            print(z_result, i_result, p_result)
            return z_result, i_result, p_result

        except Exception as e:
            self.update_status(f"Ошибка при обработке {os.path.basename(filepath)}: {str(e)}")
            return None, None


    def split_z_result(self, z_result):
        parts = ['-', '-', '-', '-']

        # Обработка первой и второй частей: до "место нахождения" и до "ОГРН"
        m_mesto = re.search(r'место нахождения:', z_result)
        if m_mesto:
            parts[0] = z_result[:m_mesto.start()].strip()
            remaining = z_result[m_mesto.end():]
            m_ogrn = re.search(r'ОГРН', remaining)
            if m_ogrn:
                parts[1] = remaining[:m_ogrn.start()].strip()
            else:
                parts[1] = '-'
        else:
            parts[0] = z_result.strip()
            parts[1] = '-'

        # Обработка третьей части: номер телефона
        tel_match = re.search(r'тел\.?\s*:?\s*([+0-9()\s-]+)', z_result, re.IGNORECASE)
        if tel_match:
            tel_str = tel_match.group(1)
            digits = re.sub(r'\D', '', tel_str)
            if len(digits) == 11:
                parts[2] = digits

        # Обработка четвертой части: email
        email_match = re.search(r'e-mail\s*:?\s*([^\s;]+?\.(?:ru|com))\b', z_result, re.IGNORECASE)
        if email_match:
            parts[3] = email_match.group(1)

        return parts


    def split_i_result(self, i_result):
        parts = ['-', '-', '-']  # [до ОГРН/адреса, адрес до телефона, телефон]

        # 1. Определяем первую часть (до ОГРН или до адреса)
        ogrn_match = re.search(r'ОГРН', i_result)
        address_match = re.search(r'адрес места осуществления деятельности(?: по изготовлению продукции)?:', i_result)

        if ogrn_match and (not address_match or ogrn_match.start() < address_match.start()):
            parts[0] = i_result[:ogrn_match.start()].strip()
        elif address_match:
            parts[0] = i_result[:address_match.start()].strip()

        # 2. Извлекаем адрес (от адресной части до телефона)
        if address_match:
            tel_match = re.search(r'тел\.?\s*:?\s*', i_result[address_match.end():], re.IGNORECASE)
            if tel_match:
                address_part = i_result[address_match.end():address_match.end() + tel_match.start()].strip(' ;')
                parts[1] = address_part
            else:
                # Если телефон не найден, берем до конца строки
                parts[1] = i_result[address_match.end():].strip(' ;')

        # 3. Извлекаем телефон
        tel_match = re.search(r'тел\.?\s*:?\s*([+0-9()\s-]+)', i_result, re.IGNORECASE)
        if tel_match:
            clean_number = re.sub(r'\D', '', tel_match.group(1))
            if len(clean_number) == 11:
                parts[2] = clean_number

        return parts


    def process_files(self, action):
        self.update_status("Начало обработки...")
        self.progress['maximum'] = len(self.file_paths)
        self.progress['value'] = 0

        try:
            if action == "new":
                wb = openpyxl.Workbook()
                ws = wb.active
                # Добавляем заголовки
                headers = ["Номер сертификата", "Заявитель", "Заявитель место нахождения", "Заявитель телефон", "Заявитель e-mail",
                          "Изготовитель", "Изготовитель адрес", "Изготовитель телефон"]
                for col_idx, header in enumerate(headers, start=1):
                    ws.cell(row=1, column=col_idx, value=header)
                row = 2
                save_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel Files", "*.xlsx")],
                    title="Сохранить новую таблицу как"
                )
            else:
                save_path = filedialog.askopenfilename(
                    filetypes=[("Excel Files", "*.xlsx")],
                    title="Выберите существующий файл Excel"
                )
                if not save_path:
                    self.update_status("Обработка отменена")
                    return
                wb = openpyxl.load_workbook(save_path)
                ws = wb.active
                row = ws.max_row + 1

            if not save_path:
                self.update_status("Обработка отменена")
                return

            for i, filepath in enumerate(self.file_paths):
                z_data, i_data, p_data = self.extract_data_from_word(filepath)

                if z_data is not None and i_data is not None:
                    z_parts = self.split_z_result(z_data)
                    i_parts = self.split_i_result(i_data)
                    all_data = p_data + z_parts + i_parts

                    # Запись данных в соответствующие столбцы
                    for col_idx, data in enumerate(all_data, start=1):
                        ws.cell(row=row, column=col_idx, value=data)
                        ws.cell(row=row, column=col_idx).alignment = Alignment(
                            wrap_text=True,
                            vertical='top'
                        )

                    row += 1

                self.update_progress(i + 1)
                self.update_status(f"Обработан файл {i + 1}/{len(self.file_paths)}")

            # Настройка ширины столбцов
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                ws.column_dimensions[col].width = 30

            wb.save(save_path)
            self.update_status(f"Данные сохранены в {save_path}")
            messagebox.showinfo("Готово", f"Данные успешно сохранены в:\n{save_path}")

        except Exception as e:
            self.update_status(f"Ошибка: {str(e)}")
            messagebox.showerror("Ошибка", f"Произошла ошибка:\n{str(e)}")
        finally:
            self.progress['value'] = 0


if __name__ == "__main__":
    app = WordToExcelConverter()
    app.mainloop()