
import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import openpyxl
import os
from datetime import datetime
import threading
from tkinterdnd2 import DND_FILES, TkinterDnD
import shutil

# Путь к файлу-шаблону
TEMPLATE_FILE_PATH = rf'\\192.168.34.9\линвит\ПОЛЬЗОВАТЕЛИ\USER49\Реестры применения\Реестр применения СИ 2024 Образец.xlsx'

# Функция для форматирования даты
def format_date(date_value):
    if isinstance(date_value, datetime):
        return date_value.strftime("%d.%m.%Y")
    else:
        return str(date_value or "")


def process_files(action):
    """
    Обрабатывает выбранные файлы, используя заранее известный индекс листа.
    """
    global cell_value1, cell_value2, cell_value3, cell_value4, cell_value4_1, cell_value5, cell_value6, cell_value7, cell_value8, cell_value10, cell_value10_1, cell_value9
    update_status("Начало обработки файлов...")  # Обновление статуса

    if not file_paths:
        messagebox.showinfo("Внимание", "Файлы не выбраны.")
        update_status("Ожидание выбора файлов.")
        return

    if action == "existing":
        save_path = filedialog.askopenfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not save_path:
            update_status("Обработка прервана пользователем.")
            return

        try:
            output_workbook = openpyxl.load_workbook(save_path)
            output_sheet = output_workbook.worksheets[0]
            output_sheet1 = output_workbook.worksheets[1]
        except FileNotFoundError:
            messagebox.showerror("Ошибка", f"Файл не найден: {save_path}")
            update_status(f"Ошибка: Файл не найден - {save_path}")
            return
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при открытии файла {save_path}: {e}")
            update_status(f"Ошибка: Не удалось открыть файл - {save_path}. Ошибка: {e}")
            return

    elif action == "new":
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not save_path:
            update_status("Обработка прервана пользователем.")
            return

        # Копируем файл шаблона
        try:
            shutil.copy(TEMPLATE_FILE_PATH, save_path)
            output_workbook = openpyxl.load_workbook(save_path)
            output_sheet = output_workbook.worksheets[0]
            output_sheet1 = output_workbook.worksheets[1]
        except FileNotFoundError:
            messagebox.showerror("Ошибка", f"Файл шаблона не найден: {TEMPLATE_FILE_PATH}")
            update_status(f"Ошибка: Файл шаблона не найден - {TEMPLATE_FILE_PATH}")
            return
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при копировании шаблона: {e}")
            update_status(f"Ошибка: Не удалось скопировать шаблон - {TEMPLATE_FILE_PATH}. Ошибка: {e}")
            return

    else:
        messagebox.showerror("Ошибка", "Некорректное действие.")
        update_status("Ошибка: Некорректное действие.")
        return

    row_index = output_sheet.max_row + 1
    row_index1 = output_sheet1.max_row + 1
    total_files = len(file_paths)
    progress_bar['maximum'] = total_files
    progress_bar['value'] = 0

    for i, file_path in enumerate(file_paths):
        try:
            update_status(f"Обработка файла {i+1} из {total_files}: {os.path.basename(file_path)}")

            workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            worksheets = workbook.sheetnames

            sheet1 = workbook['Титул']
            sheet2 = workbook['Протокол']
            sheet3 = workbook['Записи']

            for i in range(0, len(worksheets)):
                if worksheets[i] == 'Протокол-3пр':
                    sheet2 = workbook['Протокол-3пр']
                    sheet3 = workbook['Записи-3пр']

            Check1 = str(sheet2['AG34'].value)
            Check2 = str(sheet2['AG30'].value)
            Check3 = str(sheet2['AG32'].value)
            Check4 = str(sheet2['AG33'].value)

            if 'Тип СИ' in Check4:

                cell_value1 = sheet2['BE34'].value               # Заводской номер 1
                cell_value2 = sheet2['AG34'].value               # Тип СИ 1
                cell_value3 = sheet2['R21'].value                # Место нахождения СИ (регион)
                cell_value4 = sheet1['A32'].value                # Наименование организации, владельца пунктов контроля КЭ
                cell_value4_1 = sheet1['A33'].value
                if cell_value4_1 is None:
                    cell_value4_1 = ''
                cell_value5 = format_date(sheet2['M25'].value)   # Дата передачи СИ (начало испытаний)
                cell_value6 = format_date(sheet2['M26'].value)   # Дата возврата СИ (окончание испытаний)
                cell_value7 = sheet3['BZ37'].value               # Ответственный за прием / возврат СИ (измерения провел)
                cell_value8 = sheet2['AS19'].value               # Место установки
                cell_value9 = sheet1['BC26'].value               # Номер протокола
                cell_value10 = sheet2['AG35'].value              # Тип СИ 2
                cell_value10_1 = sheet2['BE35'].value            # Заводской номер 2

            if 'Тип СИ' in Check3:

                cell_value1 = sheet2['BE33'].value
                cell_value2 = sheet2['AG33'].value
                cell_value3 = sheet2['A19'].value
                cell_value4 = sheet1['A35'].value
                cell_value4_1 = sheet1['A36'].value
                cell_value5 = format_date(sheet2['M24'].value)
                cell_value6 = format_date(sheet2['M25'].value)
                cell_value7 = sheet3['BZ37'].value
                cell_value8 = sheet2['I17'].value
                cell_value9 = sheet1['BC29'].value
                cell_value10 = sheet2['AG34'].value
                cell_value10_1 = sheet2['BE34'].value

            if 'Тип СИ' in Check2:

                cell_value1 = sheet2['BE31'].value
                cell_value2 = sheet2['AG31'].value
                cell_value3 = sheet2['A19'].value
                cell_value4 = sheet1['A35'].value
                cell_value4_1 = sheet1['A36'].value
                cell_value5 = format_date(sheet2['M22'].value)
                cell_value6 = format_date(sheet2['M23'].value)
                cell_value7 = sheet3['BZ37'].value
                cell_value8 = sheet2['I18'].value
                cell_value9 = sheet1['BC29'].value
                cell_value10 = sheet2['AG32'].value
                cell_value10_1 = sheet2['BE32'].value

            elif 'Тип СИ' in Check1:

                cell_value1 = sheet2['BE35'].value               # Заводской номер 1
                cell_value2 = sheet2['AG35'].value               # Тип СИ 1
                cell_value3 = sheet2['R21'].value                # Место нахождения СИ (регион)
                cell_value4 = sheet1['A35'].value                # Наименование организации, владельца пунктов контроля КЭ
                cell_value4_1 = sheet1['A36'].value
                if cell_value4_1 is None:
                    cell_value4_1 = ''
                cell_value5 = format_date(sheet2['M26'].value)   # Дата передачи СИ (начало испытаний)
                cell_value6 = format_date(sheet2['M27'].value)   # Дата возврата СИ (окончание испытаний)
                cell_value7 = sheet3['BZ37'].value               # Ответственный за прием / возврат СИ (измерения провел)
                cell_value8 = sheet2['AS19'].value               # Место установки
                cell_value9 = sheet1['BC29'].value               # Номер протокола
                cell_value10 = sheet2['AG36'].value              # Тип СИ 2
                cell_value10_1 = sheet2['BE36'].value            # Заводской номер 2

            PorNum0 = cell_value9.split('/')
            PorNum = PorNum0[0] + ',' + PorNum0[1]

            # Запись данных в первый лист
            output_sheet[f'C{row_index}'] = cell_value1
            output_sheet[f'D{row_index}'] = cell_value2
            output_sheet[f'F{row_index}'] = cell_value3
            if cell_value4_1 is None:
                output_sheet[f'G{row_index}'] = str(cell_value4)
            else:
                output_sheet[f'G{row_index}'] = str(cell_value4) + ' ' + str(cell_value4_1)
            output_sheet[f'H{row_index}'] = cell_value5
            output_sheet[f'I{row_index}'] = cell_value6
            output_sheet[f'J{row_index}'] = cell_value7
            output_sheet[f'O{row_index}'] = cell_value8
            output_sheet[f'P{row_index}'] = cell_value9
            output_sheet[f'Q{row_index}'] = str(cell_value10) + ' Зав.№: ' + str(cell_value10_1)
            output_sheet[f'S{row_index}'] = PorNum

            # Запись данных во второй лист
            output_sheet1[f'C{row_index1}'] = cell_value10_1
            output_sheet1[f'D{row_index1}'] = cell_value10
            output_sheet1[f'F{row_index1}'] = cell_value3
            if cell_value4_1 is None:
                output_sheet1[f'G{row_index1}'] = str(cell_value4)
            else:
                output_sheet1[f'G{row_index1}'] = str(cell_value4) + ' ' + str(cell_value4_1)
            output_sheet1[f'H{row_index1}'] = cell_value5
            output_sheet1[f'I{row_index1}'] = cell_value6
            output_sheet1[f'J{row_index1}'] = cell_value7
            output_sheet1[f'O{row_index1}'] = cell_value8
            output_sheet1[f'P{row_index1}'] = cell_value9
            output_sheet1[f'Q{row_index1}'] = str(cell_value2) + ' Зав.№: ' + str(cell_value1)
            output_sheet1[f'S{row_index1}'] = PorNum

            row_index += 1
            row_index1 += 1


        except FileNotFoundError:
            messagebox.showerror("Ошибка", f"Файл не найден: {file_path}")
            update_status(f"Ошибка: Файл не найден - {os.path.basename(file_path)}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при обработке файла {os.path.basename(file_path)}: {e}")
            update_status(f"Ошибка при обработке {os.path.basename(file_path)}: {e}")

        # Обновление прогресс-бара (основной поток)
        root.after(0, update_progress)

    try:
        update_status("Сохранение файла...")

        output_workbook.save(save_path)

        messagebox.showinfo("Успех", f"Данные успешно записаны в файл: {save_path}")
        update_status(f"Данные успешно записаны в файл: {save_path}")

    except Exception as e:
        messagebox.showerror("Ошибка", f"Ошибка при сохранении файла: {e}")
        update_status(f"Ошибка при сохранении файла: {e}")
    finally:
        output_workbook.close()
        update_status("Обработка завершена.")

    root.after(0, processing_complete)  # Сообщение о завершении


def update_progress():
    """Обновляет значение прогресс-бара."""
    progress_bar['value'] += 1


def processing_complete():
    """Сбрасывает прогресс-бар после завершения."""
    progress_bar['value'] = 0

def update_status(message):
    """Обновляет текст в метке статуса."""
    status_label.config(text=message)
    root.update_idletasks()  # Обновление интерфейса для отображения сообщения


def choose_files():
    """Выбор файлов."""
    global file_paths
    new_file_paths = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx")])
    if new_file_paths:
        file_paths.extend(new_file_paths)
        update_file_list()
    update_status("Файлы выбраны.")


def delete_selected_file():
    """Удаление выбранного файла."""
    try:
        selected_index = file_list.curselection()[0]
        del file_paths[selected_index]
        update_file_list()
    except IndexError:
        messagebox.showinfo("Внимание", "Файл не выбран для удаления.")
    update_status("Файл удален из списка.")


def clear_file_list():
    """Очистка списка файлов."""
    global file_paths
    file_paths = []
    update_file_list()
    update_status("Список файлов очищен.")


def update_file_list():
    """Обновление списка файлов."""
    file_list.delete(0, tk.END)
    for file_path in file_paths:
        file_list.insert(tk.END, os.path.basename(file_path))


def drop(event):
    """Обработчик события Drop."""
    global file_paths
    files = root.tk.splitlist(event.data)
    file_paths.extend(files)
    update_file_list()
    update_status("Файлы добавлены перетаскиванием.")


def start_processing_thread(action):
    """Запуск обработки в отдельном потоке."""
    thread = threading.Thread(target=process_files, args=(action,))
    thread.start()





# == UI Setup ==
root = TkinterDnD.Tk()  # Инициализируем TkinterDnD
root.title("Excel File Transfer")

# Стиль ttk
style = ttk.Style()
style.configure("TButton", padding=5, relief="raised")
style.configure("TLabel", padding=5)

# Рамка для элементов
frame = ttk.Frame(root, padding=10)
frame.pack(expand=True, fill="both")

# Кнопка выбора файлов
choose_button = ttk.Button(frame, text="Выберите файлы или перетащите их в область ниже", command=choose_files, style="TButton")
choose_button.pack(pady=10)

file_list = tk.Listbox(frame, width=60, height=10, selectmode=tk.SINGLE)
file_list.pack(pady=5, expand=True, fill="both")

# Привязка события Drop к Listbox
file_list.drop_target_register(DND_FILES)
file_list.dnd_bind('<<Drop>>', drop)


# Фрейм для кнопок "Удалить" и "Очистить"
button_frame = ttk.Frame(frame)
button_frame.pack(pady=5)

# Кнопки управления списком
delete_button = ttk.Button(button_frame, text="Удалить выбранный файл", command=delete_selected_file, style="TButton")
delete_button.pack(side="left", padx=5)  # side="left" для размещения в ряд

clear_button = ttk.Button(button_frame, text="Очистить список файлов", command=clear_file_list, style="TButton")
clear_button.pack(side="left", padx=5)  # side="left" для размещения в ряд

# Фрейм для кнопок действий
action_button_frame = ttk.Frame(frame)
action_button_frame.pack(pady=5, fill="x")
button_width = 13  # Минимальная ширина, можно настроить

# Кнопки действий
new_button = ttk.Button(action_button_frame, text="Создать новую таблицу", command=lambda: start_processing_thread("new"), style="TButton")
new_button.pack(expand=True, fill="x")
new_button.config(width=button_width)

existing_button = ttk.Button(action_button_frame, text="Добавить в существующую таблицу", command=lambda: start_processing_thread("existing"), style="TButton")
existing_button.pack(expand=True, fill="x")
existing_button.config(width=button_width)

# Прогресс-бар
progress_bar = ttk.Progressbar(frame, orient="horizontal", length=300, mode="determinate")
progress_bar.pack(pady=10)

# Метка статуса
status_label = ttk.Label(frame, text="Ожидание выбора файлов.", anchor="center")
status_label.pack(pady=5, fill="x")


# Переменная списка файлов
file_paths = []

# Инициализация статуса
update_status("Ожидание выбора файлов.")

# Запуск UI
root.mainloop()
