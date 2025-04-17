import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# ПУТЬ К КОРНЕВОЙ ПАПКЕ
INpath = rf'\\192.168.34.9\линвит\ПОЛЬЗОВАТЕЛИ\USER49\!Сертификаты_2024'

# Названия папок для обработки
processing_folders = [
    ('0. СИ', [
        '0 Заявка и приложение',
        '1 Распоряжение по заявке',
        '2 Решение по заявке',
        '3 Заключения по ОМД и ТД',
        '4 Акт выбора ПК',
        '5 Протоколы СИ',
        '6 Заключение СИ',
        '7 Программа проверки произ',
        '8 Акт ПП',
        '9 Распоряжение на анализ',
        '10 Решение о выдаче',
        '11 Сертификат',
        '12 Доп.материалы'
    ]),
    ('1. ИК-1', [
        '0 Распоряжение',
        '1 Письмо-уведомление',
        '2 Программа ИК',
        '3 Программа проверки произ',
        '4 Акт выбора ПК',
        '5 Акт проверки производства',
        '6 Протоколы ИК',
        '7 Акт по результатам ИК',
        '8 Распоряжение на анализ',
        '9 Решение по ИК',
        '10 Доп. материалы'
    ]),
    ('2. ИК-2', [
        '0 Распоряжение',
        '1 Письмо-уведомление',
        '2 Программа ИК',
        '3 Программа проверки произ',
        '4 Акт выбора ПК',
        '5 Акт проверки производства',
        '6 Протоколы ИК',
        '7 Акт по результатам ИК',
        '8 Распоряжение на анализ',
        '9 Решение по ИК',
        '10 Доп. материалы'
    ])
]

# Создаем новую книгу Excel
wb = Workbook()
# Удаляем лист по умолчанию, если он есть
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Создаем листы для каждого типа данных
for folder, headers in processing_folders:
    ws = wb.create_sheet(title=folder.split('. ')[1])  # Используем только часть после цифры (СИ, ИК-1, ИК-2)

    # Устанавливаем ширину столбцов
    ws.column_dimensions['A'].width = 50  # Столбец A (Название папки)
    for col in range(2, len(headers) + 2):  # Остальные столбцы
        ws.column_dimensions[chr(64 + col)].width = 15

    # Записываем заголовки в первую строку
    headers_with_name = ['Название папки'] + headers

    # Устанавливаем высоту строки заголовка в два раза больше
    ws.row_dimensions[1].height = 30

    for col_num, header in enumerate(headers_with_name, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        # Включаем перенос текста и выравнивание по центру
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

# Названия папок сертификатов
folder_names = os.listdir(INpath)

for folder_name in folder_names:
    for si_name, headers in processing_folders:
        folder_path = os.path.join(INpath, folder_name, si_name)
        sheet_name = si_name.split('. ')[1]
        ws = wb[sheet_name]

        try:
            subfolder_names = os.listdir(folder_path)

            # Удаляем Thumbs.db, если он есть
            if 'Thumbs.db' in subfolder_names:
                thumbs_path = os.path.join(folder_path, 'Thumbs.db')
                os.remove(thumbs_path)
                subfolder_names.remove('Thumbs.db')
                print(f"Удален файл Thumbs.db в {folder_path}")

            # Фильтруем только нужные файлы (игнорируем нечисловые имена)
            filtered_subfolders = []
            for name in subfolder_names:
                try:
                    # Проверяем, начинается ли имя с числа (например, "1 Название")
                    int(name.split()[0])
                    filtered_subfolders.append(name)
                except (ValueError, IndexError):
                    continue  # Пропускаем файлы, не соответствующие формату

            # Сортируем по числовому префиксу
            sorted_subfolder_names = sorted(filtered_subfolders, key=lambda s: int(s.split()[0]))

            result = []
            result.append(folder_name)  # Добавляем имя папки в начало
            for s in sorted_subfolder_names:
                if "+—" in s:
                    result.append("+-")
                elif "+" in s:
                    result.append("+")
                elif "—" in s:
                    result.append("-")

            print(f"{si_name}: {result}")

            # Находим следующую свободную строку
            row_num = ws.max_row + 1 if ws.max_row > 1 else 2

            # Записываем результат в Excel
            for col_num, value in enumerate(result, start=1):
                ws.cell(row=row_num, column=col_num, value=value)

        except FileNotFoundError:
            print(f"Папка не найдена: {folder_path}")
            continue

# Сохраняем файл Excel
output_path = os.path.join(INpath, "results.xlsx")
wb.save(output_path)
print(f"Результаты сохранены в: {output_path}")